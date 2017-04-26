# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 17:19:54 2017

@author: User2
"""

'''
Analisis
El plan es leer cada linea del archivo. Cabecera, pie, y detalles.
Para la cabecera lleno una primera tabla de cabecera en hoja1. Se la posicion. 
Para detalles, lleno otra hoja2. 
Para pie, lleno otra hoja3
Las columnas de las tablas puedo ponerlas yo.

Diseño
Leo linea, la mando a funcion.
Funcion la parto en pedacitos que van a cada columna.
Para partirla necesito saber cuan largo es. Nuevo modulo con tupla.

Defino una funcion que genere columnas (saca info de modulo)
Defino otra funcion para llenar tabla (llega string y splitea)

'''

#librerias
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
import structFile
#defino workbook
wb = Workbook()
#defino hojas de trabajo
hojaCabe = wb.active #Consigo la primera generada por default
hojaCabe.title = "Cabecera" #le cambio el nombre
hojaDeta = wb.create_sheet("Detalle") 
hojaPie = wb.create_sheet("Pie") 
#Muestro mis hojas (deberian ser 3)
print(wb.sheetnames)
#Estilo
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

def generarArchivo():
    #declaro archivo para debug
    archivo = 'pagomiscuentas.txt'
    #formatos
    formatoCabe = structFile.formatoCabe    
    formatoDeta = structFile.formatoDeta
    formatoPie =  structFile.formatoPie
    #print formatoCabe[2][0]
    if archivo == "":
        archivo = raw_input("Ingrese el nombre del archivo incluido extension: ")    
    
    with open(archivo, "r+") as f:
        print f.name + " abierto correctamente!\n"
        cantiRenglones = sum(1 for line in f)
        #print cantiRenglones
        f.seek(0) #dejo el cursor en el inicio de nuevo
        insertarColumnas(formatoCabe,hojaCabe)
        insertarColumnas(formatoDeta,hojaDeta)
        insertarColumnas(formatoPie,hojaPie)
        
        for i,registro in enumerate(f):
            if i == 0:
                insertarDatos(0,registro,formatoCabe,hojaCabe)
            elif i == cantiRenglones-1: #-1 porque i empieza en 0
                insertarDatos(0,registro,formatoPie,hojaPie)                
            else:
                insertarDatos(i-1,registro,formatoDeta,hojaDeta)
    
    wb.save('PygoMisCuentas.xlsx')
    print "Archivo correctamente generado!"
    
def insertarColumnas(formato,hoja):
    base = 2 #empiezo en la segunda linea-columna
    for i,j in enumerate(range(base,len(formato)+base)) :#j es parte del formato
        hoja.cell(row=base, column = j).value = formato[i][0] #i va de 0->X  
        hoja.cell(row=base, column = j).border = thin_border
        #print formato[i][0]
#    for i in range(base,len(formato)+base) :
#        print hoja.cell(row=base, column = i).value
#            
def insertarDatos(nume,linea,formato,hoja):
    base = 3 #empiezo en la tercera linea-columna
    for i,j in enumerate(range(base,len(formato)+base)) :#j es parte del formato
        #print formato[i][1]
        if formato[i][1] == "Decimal":
            hoja.cell(row=nume+base, column = j-1).number_format = '0.00'
            hoja.cell(row=nume+base, column = j-1).value = convertirADecimal(procesarLinea(linea,formato[i][3])) #i va de 0->X              
        else:
            hoja.cell(row=nume+base, column = j-1).value = procesarLinea(linea,formato[i][3]) #i va de 0->X  
def procesarLinea(registro,bound):
    limites = bound.split('-')
    #print limites
    desde = int(limites[0])
    hasta = int(limites[1])
    #print registro[desde-1:hasta]
    return registro[desde-1:hasta]
    
def convertirADecimal(numero):  
    if numero <> "":       
        entero = int(numero[:-2])
        decimal = int(numero[-2:])
        decimal *= 0.01 
        return entero + decimal
    else:
        return 0
        
generarArchivo()
